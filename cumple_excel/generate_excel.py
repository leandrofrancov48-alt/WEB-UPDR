import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

def create_birthday_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Invitados"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Define styles
    font_family = "Segoe UI"
    
    # Colors (HEX codes without # for openpyxl)
    NAVY_HEADER = "1A365D"      # Dark Navy
    WHITE = "FFFFFF"
    SOFT_BLUE = "F0F4F8"       # Light blue for alternating rows
    BORDER_GREY = "D2D6DC"     # Light grey for standard borders
    CARD_BG = "F7FAFC"         # Very light grey for dashboard cards
    CARD_BORDER = "CBD5E0"     # Card border
    
    # Text colors
    TEXT_NAVY = "1A365D"
    TEXT_MUTED = "718096"
    
    # Conditional formatting colors
    CONFIRMED_FILL = "D4EDDA"   # Light Green
    CONFIRMED_TEXT = "155724"   # Dark Green
    PENDING_FILL = "FFF3CD"     # Light Yellow
    PENDING_TEXT = "856404"     # Dark Yellow
    DECLINED_FILL = "F8D7DA"    # Light Red
    DECLINED_TEXT = "721C24"    # Dark Red
    NA_FILL = "EDF2F7"          # Soft Grey
    NA_TEXT = "718096"          # Dark Grey
    
    # Fonts
    title_font = Font(name=font_family, size=18, bold=True, color=TEXT_NAVY)
    subtitle_font = Font(name=font_family, size=10, italic=True, color=TEXT_MUTED)
    header_font = Font(name=font_family, size=11, bold=True, color=WHITE)
    data_font = Font(name=font_family, size=11, color="000000")
    bold_data_font = Font(name=font_family, size=11, bold=True, color="000000")
    card_title_font = Font(name=font_family, size=10, bold=True, color=TEXT_MUTED)
    card_value_font = Font(name=font_family, size=20, bold=True, color=TEXT_NAVY)
    
    # Fills
    header_fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    alt_row_fill = PatternFill(start_color=SOFT_BLUE, end_color=SOFT_BLUE, fill_type="solid")
    card_fill = PatternFill(start_color=CARD_BG, end_color=CARD_BG, fill_type="solid")
    
    # Alignments
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    title_align = Alignment(horizontal="left", vertical="center")
    
    # Borders
    thin_border_side = Side(border_style="thin", color=BORDER_GREY)
    thick_bottom_side = Side(border_style="medium", color=NAVY_HEADER)
    double_bottom_side = Side(border_style="double", color="000000")
    
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    total_border = Border(top=thin_border_side, bottom=double_bottom_side)
    
    card_border_side = Side(border_style="thin", color=CARD_BORDER)
    card_border = Border(left=card_border_side, right=card_border_side, top=card_border_side, bottom=card_border_side)

    # Set Column Widths
    column_widths = {
        'A': 3,   # Spacer
        'B': 6,   # N°
        'C': 20,  # Nombre
        'D': 16,  # Cena
        'E': 16,  # Fiesta
        'F': 25,  # Bebida preferida
        'G': 30,  # Restricciones / Notas
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Set Row Heights
    ws.row_dimensions[2].height = 30  # Title row
    ws.row_dimensions[3].height = 18  # Subtitle row
    ws.row_dimensions[5].height = 20  # Card Title row
    ws.row_dimensions[6].height = 32  # Card Value row
    ws.row_dimensions[10].height = 28 # Table Header row

    # --- Title Section ---
    ws['B2'] = "LISTA DE INVITADOS - CUMPLEAÑOS 🎉"
    ws['B2'].font = title_font
    ws['B2'].alignment = title_align
    ws.merge_cells('B2:G2')

    ws['B3'] = "Planificación de cena, fiesta y asistencia total"
    ws['B3'].font = subtitle_font
    ws['B3'].alignment = title_align
    ws.merge_cells('B3:G3')

    # --- Dashboard Cards (Summary Section) ---
    # Card 1: Cena Confirmados
    ws.merge_cells('B5:C5')
    ws['B5'] = "CONFIRMADOS CENA 🍽️"
    ws['B5'].font = card_title_font
    ws['B5'].alignment = center_align
    
    ws.merge_cells('B6:C6')
    # Count of 'Confirmado' in Cena column (D11:D30)
    ws['B6'] = '=COUNTIF(D11:D30, "Confirmado") & " / " & COUNTIFS(D11:D30, "<>N/A")'
    ws['B6'].font = card_value_font
    ws['B6'].alignment = center_align
    
    # Card 2: Fiesta Confirmados
    ws.merge_cells('D5:E5')
    ws['D5'] = "CONFIRMADOS FIESTA 🥳"
    ws['D5'].font = card_title_font
    ws['D5'].alignment = center_align
    
    ws.merge_cells('D6:E6')
    # Count of 'Confirmado' in Fiesta column (E11:E30)
    ws['D6'] = '=COUNTIF(E11:E30, "Confirmado") & " / " & COUNTA(C11:C30)'
    ws['D6'].font = card_value_font
    ws['D6'].alignment = center_align

    # Card 3: Total Confirmados (Cena + Fiesta únicos)
    ws.merge_cells('F5:G5')
    ws['F5'] = "TOTAL INVITADOS UNICOS 👥"
    ws['F5'].font = card_title_font
    ws['F5'].alignment = center_align
    
    ws.merge_cells('F6:G6')
    # Total unique count of guests
    ws['F6'] = '=COUNTA(C11:C30)'
    ws['F6'].font = card_value_font
    ws['F6'].alignment = center_align

    # Style Cards
    for col_idx in ['B', 'C', 'D', 'E', 'F', 'G']:
        for row_idx in [5, 6]:
            cell = ws[f"{col_idx}{row_idx}"]
            cell.fill = card_fill
            
            # Apply border around cards (manually drawing borders on cells)
            left_side = card_border_side if col_idx in ['B', 'D', 'F'] else None
            right_side = card_border_side if col_idx in ['C', 'E', 'G'] else None
            top_side = card_border_side if row_idx == 5 else None
            bottom_side = card_border_side if row_idx == 6 else None
            
            cell.border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)

    # --- Guest List Table ---
    # Headers
    headers = ["N°", "Nombre", "Cena 🍽️", "Fiesta 🥳", "Bebida Preferida 🍾", "Notas / Alergias 📝"]
    for i, header in enumerate(headers):
        col_letter = chr(ord('B') + i)
        cell = ws[f"{col_letter}10"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if col_letter in ['B', 'D', 'E'] else left_align
        cell.border = header_border

    # Guest Data
    # Cena guests (15)
    cena_guests = [
        "Lean", "Ro", "Ivan", "Fran", "Marco", "Sinpi", "Nake", "Nqh", 
        "Agape", "Agustina", "Ani", "Diego", "Charra 1", "Charra 2", "Valen Local"
    ]
    # Voley guests (5)
    voley_guests = ["Voley 1", "Voley 2", "Voley 3", "Voley 4", "Voley 5"]
    
    all_guests = cena_guests + voley_guests

    start_row = 11
    for idx, name in enumerate(all_guests):
        row = start_row + idx
        ws.row_dimensions[row].height = 22
        
        # N°
        ws[f"B{row}"] = idx + 1
        ws[f"B{row}"].font = bold_data_font
        ws[f"B{row}"].alignment = center_align
        
        # Nombre
        ws[f"C{row}"] = name
        ws[f"C{row}"].font = data_font
        ws[f"C{row}"].alignment = left_align
        
        # Cena status
        # If the guest is from Volleyball, they aren't invited to dinner -> "N/A"
        if name in voley_guests:
            ws[f"D{row}"] = "N/A"
        else:
            ws[f"D{row}"] = "Pendiente"
        ws[f"D{row}"].font = data_font
        ws[f"D{row}"].alignment = center_align
        
        # Fiesta status
        ws[f"E{row}"] = "Pendiente"
        ws[f"E{row}"].font = data_font
        ws[f"E{row}"].alignment = center_align
        
        # Bebida Preferida
        ws[f"F{row}"] = ""
        ws[f"F{row}"].font = data_font
        ws[f"F{row}"].alignment = left_align
        
        # Notas / Alergias
        ws[f"G{row}"] = ""
        ws[f"G{row}"].font = data_font
        ws[f"G{row}"].alignment = left_align
        
        # Apply standard borders and alternating background row color (for white/soft blue)
        row_fill = alt_row_fill if idx % 2 == 1 else PatternFill(fill_type=None)
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
            cell = ws[f"{col_letter}{row}"]
            cell.border = data_border
            if row_fill.fill_type:
                cell.fill = row_fill

    end_row = start_row + len(all_guests) - 1

    # --- Data Validation Dropdowns ---
    # Validation for status (Confirmado, Pendiente, Rechazado)
    status_validation = DataValidation(
        type="list", 
        formula1='"Confirmado,Pendiente,Rechazado"', 
        allow_blank=True
    )
    status_validation.error ='El estado debe ser Confirmado, Pendiente o Rechazado'
    status_validation.errorTitle = 'Entrada inválida'
    status_validation.prompt = 'Selecciona un estado de la lista'
    status_validation.promptTitle = 'Estado de Asistencia'
    
    ws.add_data_validation(status_validation)
    
    # Add validation range to Cena (only for non-voley guests) and Fiesta (all guests)
    status_validation.add(f"D{start_row}:D{start_row + len(cena_guests) - 1}")
    status_validation.add(f"E{start_row}:E{end_row}")

    # Dropdown validation for Cena N/A cell (just so they can choose N/A or others if needed)
    na_validation = DataValidation(
        type="list",
        formula1='"Confirmado,Pendiente,Rechazado,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(na_validation)
    na_validation.add(f"D{start_row + len(cena_guests)}:D{end_row}")

    # --- Conditional Formatting Rules ---
    # Rules to color cells dynamically based on text value
    def add_conditional_formatting(ws, column_range):
        # Confirmado (Green)
        green_fill = PatternFill(start_color=CONFIRMED_FILL, end_color=CONFIRMED_FILL, fill_type="solid")
        green_font = Font(name=font_family, size=11, color=CONFIRMED_TEXT, bold=True)
        rule_green = CellIsRule(operator='equal', formula=['"Confirmado"'], fill=green_fill, font=green_font)
        ws.conditional_formatting.add(column_range, rule_green)
        
        # Pendiente (Yellow)
        yellow_fill = PatternFill(start_color=PENDING_FILL, end_color=PENDING_FILL, fill_type="solid")
        yellow_font = Font(name=font_family, size=11, color=PENDING_TEXT, bold=True)
        rule_yellow = CellIsRule(operator='equal', formula=['"Pendiente"'], fill=yellow_fill, font=yellow_font)
        ws.conditional_formatting.add(column_range, rule_yellow)
        
        # Rechazado (Red)
        red_fill = PatternFill(start_color=DECLINED_FILL, end_color=DECLINED_FILL, fill_type="solid")
        red_font = Font(name=font_family, size=11, color=DECLINED_TEXT, bold=True)
        rule_red = CellIsRule(operator='equal', formula=['"Rechazado"'], fill=red_fill, font=red_font)
        ws.conditional_formatting.add(column_range, rule_red)
        
        # N/A (Grey)
        grey_fill = PatternFill(start_color=NA_FILL, end_color=NA_FILL, fill_type="solid")
        grey_font = Font(name=font_family, size=11, color=NA_TEXT, italic=True)
        rule_grey = CellIsRule(operator='equal', formula=['"N/A"'], fill=grey_fill, font=grey_font)
        ws.conditional_formatting.add(column_range, rule_grey)

    add_conditional_formatting(ws, f"D{start_row}:D{end_row}")
    add_conditional_formatting(ws, f"E{start_row}:E{end_row}")

    # --- Save Workbook ---
    output_path = "C:\\Users\\Lean\\.gemini\\antigravity\\scratch\\cumple_excel\\Invitados_Cumple.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Excel created successfully at: {output_path}")

if __name__ == "__main__":
    create_birthday_excel()
